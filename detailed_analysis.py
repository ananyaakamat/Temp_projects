import openpyxl
from openpyxl import load_workbook
import pandas as pd
from collections import defaultdict

def analyze_conversion_details():
    """
    Provide detailed analysis of the conversion results.
    """
    file_path = r"d:\Anant\VSCodeProjects\Temp_projects\ToNumber.xlsx"

    print("Loading Excel file for detailed analysis...")
    workbook = load_workbook(file_path)

    original_sheet = workbook["RawData"]
    converted_sheet = workbook["RawData_Numbers"]

    # Define the range
    start_col = 4  # Column D
    end_col = 27   # Column AA
    start_row = 2
    last_row = min(original_sheet.max_row, converted_sheet.max_row)

    print(f"Analyzing data from row {start_row} to {last_row}, columns D to AA")
    print("=" * 80)

    # Statistics tracking
    stats = {
        'total_cells': 0,
        'original_types': defaultdict(int),
        'converted_types': defaultdict(int),
        'conversion_patterns': defaultdict(int),
        'missing_data_indicators': defaultdict(int),
        'decimal_removals': 0,
        'columns_processed': []
    }

    # Analyze each column
    for col in range(start_col, end_col + 1):
        col_letter = chr(64 + col)
        stats['columns_processed'].append(col_letter)

        print(f"\nAnalyzing Column {col_letter}:")
        print("-" * 40)

        col_stats = {
            'original_numbers': 0,
            'original_floats': 0,
            'original_integers': 0,
            'original_missing': 0,
            'converted_integers': 0,
            'converted_empty': 0,
            'decimal_cases': []
        }

        # Sample some values from this column
        sample_values = []

        for row in range(start_row, min(start_row + 10, last_row + 1)):
            stats['total_cells'] += 1

            original_value = original_sheet.cell(row=row, column=col).value
            converted_value = converted_sheet.cell(row=row, column=col).value

            # Track original types
            if original_value is None:
                stats['original_types']['None'] += 1
                col_stats['original_missing'] += 1
            elif isinstance(original_value, (int, float)):
                stats['original_types']['number'] += 1
                col_stats['original_numbers'] += 1
                if isinstance(original_value, float):
                    col_stats['original_floats'] += 1
                    if original_value != int(original_value):
                        col_stats['decimal_cases'].append((row, original_value, converted_value))
                        stats['decimal_removals'] += 1
                else:
                    col_stats['original_integers'] += 1
            else:
                str_val = str(original_value).strip()
                stats['original_types']['string'] += 1
                if str_val in ["--", " --", "-- ", " -- "]:
                    stats['missing_data_indicators'][str_val] += 1
                    col_stats['original_missing'] += 1

            # Track converted types
            if converted_value is None or converted_value == "":
                stats['converted_types']['empty'] += 1
                col_stats['converted_empty'] += 1
            elif isinstance(converted_value, int):
                stats['converted_types']['integer'] += 1
                col_stats['converted_integers'] += 1
            elif isinstance(converted_value, float):
                stats['converted_types']['float'] += 1
            else:
                stats['converted_types']['other'] += 1

            # Track conversion patterns
            orig_type = type(original_value).__name__
            conv_type = type(converted_value).__name__
            stats['conversion_patterns'][f"{orig_type} -> {conv_type}"] += 1

            # Collect sample values
            if len(sample_values) < 5:
                sample_values.append((row, original_value, converted_value))

        # Print column summary
        print(f"Numbers found: {col_stats['original_numbers']}")
        print(f"Missing data: {col_stats['original_missing']}")
        print(f"Converted to integers: {col_stats['converted_integers']}")
        print(f"Converted to empty: {col_stats['converted_empty']}")

        if col_stats['decimal_cases']:
            print(f"Decimal removals: {len(col_stats['decimal_cases'])}")
            for row, orig, conv in col_stats['decimal_cases'][:3]:  # Show first 3
                print(f"  Row {row}: {orig} -> {conv}")

        # Show sample values
        print("Sample values:")
        for row, orig, conv in sample_values:
            print(f"  Row {row}: '{orig}' -> '{conv}'")

    # Overall statistics report
    print("\n" + "=" * 80)
    print("DETAILED CONVERSION ANALYSIS")
    print("=" * 80)

    print(f"Total cells analyzed: {stats['total_cells']}")
    print(f"Columns processed: {', '.join(stats['columns_processed'])}")
    print(f"Decimal values converted: {stats['decimal_removals']}")

    print(f"\nOriginal data types:")
    for dtype, count in stats['original_types'].items():
        percentage = (count / stats['total_cells']) * 100
        print(f"  {dtype}: {count} ({percentage:.1f}%)")

    print(f"\nConverted data types:")
    for dtype, count in stats['converted_types'].items():
        percentage = (count / stats['total_cells']) * 100
        print(f"  {dtype}: {count} ({percentage:.1f}%)")

    print(f"\nMissing data indicators found:")
    for indicator, count in stats['missing_data_indicators'].items():
        print(f"  '{indicator}': {count}")

    print(f"\nConversion patterns (top 10):")
    sorted_patterns = sorted(stats['conversion_patterns'].items(), key=lambda x: x[1], reverse=True)
    for pattern, count in sorted_patterns[:10]:
        percentage = (count / stats['total_cells']) * 100
        print(f"  {pattern}: {count} ({percentage:.1f}%)")

    workbook.close()

    # Data integrity checks
    print(f"\n" + "=" * 80)
    print("DATA INTEGRITY SUMMARY")
    print("=" * 80)

    total_numbers = stats['original_types']['number']
    total_converted_integers = stats['converted_types']['integer']
    total_missing = stats['original_types']['None'] + sum(stats['missing_data_indicators'].values())
    total_converted_empty = stats['converted_types']['empty']

    print(f"✅ Numbers in original: {total_numbers}")
    print(f"✅ Integers in converted: {total_converted_integers}")
    print(f"✅ Missing in original: {total_missing}")
    print(f"✅ Empty in converted: {total_converted_empty}")
    print(f"✅ Decimals removed: {stats['decimal_removals']}")

    # Check data preservation
    if total_converted_integers >= total_numbers * 0.95:  # Allow 5% tolerance
        print(f"\n✅ CONVERSION SUCCESS: Most numbers were properly converted to integers")
    else:
        print(f"\n⚠️  CONVERSION WARNING: Some numbers may not have been converted properly")

    if total_converted_empty >= total_missing * 0.95:  # Allow 5% tolerance
        print(f"✅ MISSING DATA SUCCESS: Missing data was properly preserved")
    else:
        print(f"⚠️  MISSING DATA WARNING: Some missing data may not have been handled properly")

if __name__ == "__main__":
    try:
        analyze_conversion_details()
        input("\nPress Enter to exit...")

    except Exception as e:
        print(f"Error during analysis: {str(e)}")
        input("Press Enter to exit...")
