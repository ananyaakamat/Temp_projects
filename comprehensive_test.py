import openpyxl
from openpyxl import load_workbook
import random

def comprehensive_test():
    """
    Comprehensive test suite for the Excel conversion validation.
    """
    print("COMPREHENSIVE CONVERSION TEST SUITE")
    print("=" * 80)

    file_path = r"d:\Anant\VSCodeProjects\Temp_projects\ToNumber.xlsx"

    # Test 1: Basic file and sheet existence
    print("1. Testing file and sheet existence...")
    try:
        workbook = load_workbook(file_path)
        assert "RawData" in workbook.sheetnames, "RawData sheet missing"
        assert "RawData_Numbers" in workbook.sheetnames, "RawData_Numbers sheet missing"
        print("   ✅ Both sheets exist")
    except Exception as e:
        print(f"   ❌ File/Sheet error: {e}")
        return False

    original_sheet = workbook["RawData"]
    converted_sheet = workbook["RawData_Numbers"]

    # Test 2: Sheet dimensions match
    print("2. Testing sheet dimensions...")
    orig_rows = original_sheet.max_row
    conv_rows = converted_sheet.max_row
    orig_cols = original_sheet.max_column
    conv_cols = converted_sheet.max_column

    if orig_rows == conv_rows and orig_cols == conv_cols:
        print(f"   ✅ Dimensions match: {orig_rows} rows x {orig_cols} columns")
    else:
        print(f"   ❌ Dimension mismatch: Original({orig_rows}x{orig_cols}) vs Converted({conv_rows}x{conv_cols})")

    # Test 3: Header preservation (row 1)
    print("3. Testing header preservation...")
    header_errors = 0
    start_col, end_col = 4, 27  # D to AA

    for col in range(1, orig_cols + 1):
        orig_header = original_sheet.cell(row=1, column=col).value
        conv_header = converted_sheet.cell(row=1, column=col).value
        if orig_header != conv_header:
            header_errors += 1
            if header_errors <= 3:  # Show first 3 errors
                print(f"   ❌ Header mismatch in column {chr(64+col)}: '{orig_header}' vs '{conv_header}'")

    if header_errors == 0:
        print("   ✅ All headers preserved correctly")
    else:
        print(f"   ❌ {header_errors} header mismatches found")

    # Test 4: Data type conversion in target range
    print("4. Testing data type conversions in range D-AA...")

    test_results = {
        'numbers_converted': 0,
        'missing_preserved': 0,
        'type_errors': 0,
        'decimal_removals': 0,
        'unexpected_changes': 0
    }

    # Random sampling for performance
    sample_rows = random.sample(range(2, min(100, orig_rows + 1)), min(20, orig_rows - 1))
    sample_cols = random.sample(range(start_col, end_col + 1), min(10, end_col - start_col + 1))

    for row in sample_rows:
        for col in sample_cols:
            orig_val = original_sheet.cell(row=row, column=col).value
            conv_val = converted_sheet.cell(row=row, column=col).value

            # Test missing data preservation
            if orig_val is None or str(orig_val).strip() in [" --", "--", "-- ", " -- ", "", "N/A", "NA"]:
                if conv_val is None or conv_val == "":
                    test_results['missing_preserved'] += 1
                else:
                    test_results['type_errors'] += 1
                    print(f"   ❌ Missing data not preserved at {chr(64+col)}{row}: '{orig_val}' -> '{conv_val}'")

            # Test number conversion
            elif orig_val is not None:
                try:
                    orig_num = float(str(orig_val))
                    expected_int = int(orig_num)

                    if orig_num != expected_int:
                        test_results['decimal_removals'] += 1

                    if isinstance(conv_val, int) and conv_val == expected_int:
                        test_results['numbers_converted'] += 1
                    elif isinstance(conv_val, float) and conv_val == expected_int:
                        test_results['numbers_converted'] += 1
                        print(f"   ⚠️ Number converted but still float type at {chr(64+col)}{row}")
                    else:
                        test_results['type_errors'] += 1
                        print(f"   ❌ Number conversion error at {chr(64+col)}{row}: '{orig_val}' -> '{conv_val}'")

                except (ValueError, TypeError):
                    # Original wasn't a number, check if it changed unexpectedly
                    if orig_val != conv_val and not (conv_val is None or conv_val == ""):
                        test_results['unexpected_changes'] += 1

    print(f"   Numbers converted correctly: {test_results['numbers_converted']}")
    print(f"   Missing data preserved: {test_results['missing_preserved']}")
    print(f"   Decimal values processed: {test_results['decimal_removals']}")
    print(f"   Type conversion errors: {test_results['type_errors']}")
    print(f"   Unexpected changes: {test_results['unexpected_changes']}")

    # Test 5: Data outside target range should be unchanged
    print("5. Testing data outside target range (columns A-C, AB+)...")
    unchanged_errors = 0

    # Test columns A, B, C (before target range)
    for col in range(1, start_col):
        for row in random.sample(range(1, min(20, orig_rows + 1)), min(5, orig_rows)):
            orig_val = original_sheet.cell(row=row, column=col).value
            conv_val = converted_sheet.cell(row=row, column=col).value
            if orig_val != conv_val:
                unchanged_errors += 1
                if unchanged_errors <= 3:
                    print(f"   ❌ Unexpected change in column {chr(64+col)}: '{orig_val}' -> '{conv_val}'")

    # Test columns after AA (if any exist)
    if orig_cols > end_col:
        for col in range(end_col + 1, min(end_col + 5, orig_cols + 1)):
            for row in random.sample(range(1, min(20, orig_rows + 1)), min(5, orig_rows)):
                orig_val = original_sheet.cell(row=row, column=col).value
                conv_val = converted_sheet.cell(row=row, column=col).value
                if orig_val != conv_val:
                    unchanged_errors += 1
                    if unchanged_errors <= 3:
                        col_name = ""
                        if col <= 26:
                            col_name = chr(64 + col)
                        else:
                            col_name = chr(64 + (col - 1) // 26) + chr(64 + ((col - 1) % 26) + 1)
                        print(f"   ❌ Unexpected change in column {col_name}: '{orig_val}' -> '{conv_val}'")

    if unchanged_errors == 0:
        print("   ✅ Data outside target range unchanged")
    else:
        print(f"   ❌ {unchanged_errors} unexpected changes outside target range")

    # Test 6: Performance and completeness check
    print("6. Testing performance and completeness...")
    target_cells = (end_col - start_col + 1) * (orig_rows - 1)  # Exclude header row
    print(f"   Target range contains {target_cells} cells to process")
    print(f"   File dimensions: {orig_rows} rows x {orig_cols} columns")

    # Test 7: Data integrity spot checks
    print("7. Performing data integrity spot checks...")

    # Check some specific patterns
    spot_check_passed = True

    # Check for any float values in converted data within target range
    float_count = 0
    for row in range(2, min(orig_rows + 1, 12)):  # Check first 10 data rows
        for col in range(start_col, end_col + 1):
            conv_val = converted_sheet.cell(row=row, column=col).value
            if isinstance(conv_val, float):
                float_count += 1
                if float_count <= 3:
                    print(f"   ⚠️ Float value found at {chr(64+col)}{row}: {conv_val}")

    if float_count == 0:
        print("   ✅ No unexpected float values found")
    else:
        print(f"   ⚠️ Found {float_count} float values (may be acceptable if they're whole numbers)")

    workbook.close()

    # Final assessment
    print("\n" + "=" * 80)
    print("FINAL TEST RESULTS")
    print("=" * 80)

    total_errors = (header_errors + test_results['type_errors'] +
                   test_results['unexpected_changes'] + unchanged_errors)

    if total_errors == 0:
        print("🎉 ALL TESTS PASSED! The conversion was successful.")
        print("\nSummary:")
        print(f"✅ Headers preserved correctly")
        print(f"✅ Numbers converted to integers: {test_results['numbers_converted']}")
        print(f"✅ Missing data preserved: {test_results['missing_preserved']}")
        print(f"✅ Decimal values processed: {test_results['decimal_removals']}")
        print(f"✅ Data outside target range unchanged")
        return True
    else:
        print(f"❌ TESTS FAILED! Found {total_errors} errors.")
        print("\nIssues found:")
        if header_errors > 0:
            print(f"❌ Header errors: {header_errors}")
        if test_results['type_errors'] > 0:
            print(f"❌ Type conversion errors: {test_results['type_errors']}")
        if test_results['unexpected_changes'] > 0:
            print(f"❌ Unexpected changes: {test_results['unexpected_changes']}")
        if unchanged_errors > 0:
            print(f"❌ Changes outside target range: {unchanged_errors}")
        return False

if __name__ == "__main__":
    try:
        success = comprehensive_test()

        print(f"\n{'='*80}")
        if success:
            print("🎉 CONVERSION VALIDATION: COMPLETE SUCCESS!")
        else:
            print("❌ CONVERSION VALIDATION: ISSUES FOUND!")
        print(f"{'='*80}")

        input("\nPress Enter to exit...")

    except Exception as e:
        print(f"❌ Error during comprehensive testing: {str(e)}")
        input("Press Enter to exit...")
