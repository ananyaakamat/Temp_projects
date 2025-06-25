import openpyxl
from openpyxl import load_workbook
import re

def is_missing_data(value):
    """Check if a value represents missing data."""
    if value is None:
        return True
    str_value = str(value).strip()
    return str_value in ["--", " --", "-- ", " -- ", "", "N/A", "NA", "n/a"]

def is_number(value):
    """Check if a value is a number."""
    if value is None:
        return False
    try:
        float(str(value))
        return True
    except (ValueError, TypeError):
        return False

def validate_conversion():
    """
    Validate the conversion results by comparing RawData and RawData_Numbers sheets.
    """
    file_path = r"d:\Anant\VSCodeProjects\Temp_projects\ToNumber.xlsx"
    
    print("Loading Excel file for validation...")
    workbook = load_workbook(file_path)
    
    # Check if both sheets exist
    if "RawData" not in workbook.sheetnames:
        print("ERROR: RawData sheet not found!")
        return False
    
    if "RawData_Numbers" not in workbook.sheetnames:
        print("ERROR: RawData_Numbers sheet not found!")
        return False
    
    original_sheet = workbook["RawData"]
    converted_sheet = workbook["RawData_Numbers"]
    
    # Define the range to validate (columns D to AA, rows 2 to last row)
    start_col = 4  # Column D
    end_col = 27   # Column AA
    start_row = 2
    
    last_row = min(original_sheet.max_row, converted_sheet.max_row)
    
    print(f"Validating conversion from row {start_row} to {last_row}, columns D to AA")
    print("=" * 70)
    
    # Validation counters
    total_cells = 0
    converted_cells = 0
    preserved_missing = 0
    preserved_headers = 0
    errors = 0
    
    # Validate headers (row 1)
    print("1. Validating headers (row 1)...")
    for col in range(start_col, end_col + 1):
        original_header = original_sheet.cell(row=1, column=col).value
        converted_header = converted_sheet.cell(row=1, column=col).value
        
        if original_header == converted_header:
            preserved_headers += 1
        else:
            print(f"   ERROR: Header mismatch in column {chr(64+col)}: '{original_header}' vs '{converted_header}'")
            errors += 1
    
    print(f"   Headers preserved: {preserved_headers}/{end_col - start_col + 1}")
    
    # Validate data conversion (rows 2 onwards)
    print("\n2. Validating data conversion...")
    
    for row in range(start_row, last_row + 1):
        for col in range(start_col, end_col + 1):
            total_cells += 1
            
            original_value = original_sheet.cell(row=row, column=col).value
            converted_value = converted_sheet.cell(row=row, column=col).value
            
            # Check if original was missing data
            if is_missing_data(original_value):
                if converted_value == "" or converted_value is None:
                    preserved_missing += 1
                else:
                    print(f"   ERROR: Row {row}, Col {chr(64+col)}: Missing data not preserved")
                    print(f"          Original: '{original_value}' -> Converted: '{converted_value}'")
                    errors += 1
            
            # Check if original was a number
            elif is_number(original_value):
                try:
                    original_num = float(str(original_value))
                    
                    # Check if converted value is an integer
                    if isinstance(converted_value, int):
                        expected_int = int(original_num)
                        if converted_value == expected_int:
                            converted_cells += 1
                        else:
                            print(f"   ERROR: Row {row}, Col {chr(64+col)}: Incorrect conversion")
                            print(f"          Original: {original_value} -> Expected: {expected_int} -> Got: {converted_value}")
                            errors += 1
                    elif isinstance(converted_value, float):
                        # Should not have decimals in converted data
                        if converted_value == int(converted_value):
                            print(f"   WARNING: Row {row}, Col {chr(64+col)}: Still has float type but correct value")
                            converted_cells += 1
                        else:
                            print(f"   ERROR: Row {row}, Col {chr(64+col)}: Still has decimals")
                            print(f"          Original: {original_value} -> Converted: {converted_value}")
                            errors += 1
                    else:
                        print(f"   ERROR: Row {row}, Col {chr(64+col)}: Number not converted properly")
                        print(f"          Original: {original_value} -> Converted: {converted_value} (type: {type(converted_value)})")
                        errors += 1
                        
                except (ValueError, TypeError):
                    print(f"   ERROR: Row {row}, Col {chr(64+col)}: Could not validate number conversion")
                    errors += 1
            
            # Progress indicator for large datasets
            if total_cells % 5000 == 0:
                print(f"   Validated {total_cells} cells...")
    
    # Generate validation report
    print("\n" + "=" * 70)
    print("VALIDATION REPORT")
    print("=" * 70)
    print(f"Total cells validated: {total_cells}")
    print(f"Headers preserved: {preserved_headers}/{end_col - start_col + 1}")
    print(f"Numbers converted: {converted_cells}")
    print(f"Missing data preserved: {preserved_missing}")
    print(f"Errors found: {errors}")
    
    # Sample data comparison
    print("\n3. Sample data comparison (first 5 rows):")
    print("-" * 70)
    print(f"{'Row':<4} {'Col':<4} {'Original':<15} {'Converted':<15} {'Status':<10}")
    print("-" * 70)
    
    sample_count = 0
    for row in range(start_row, min(start_row + 5, last_row + 1)):
        for col in range(start_col, min(start_col + 5, end_col + 1)):
            original_value = original_sheet.cell(row=row, column=col).value
            converted_value = converted_sheet.cell(row=row, column=col).value
            
            # Determine status
            if is_missing_data(original_value):
                status = "Preserved" if (converted_value == "" or converted_value is None) else "ERROR"
            elif is_number(original_value):
                try:
                    expected = int(float(str(original_value)))
                    status = "Converted" if converted_value == expected else "ERROR"
                except:
                    status = "ERROR"
            else:
                status = "Unchanged"
            
            print(f"{row:<4} {chr(64+col):<4} {str(original_value):<15} {str(converted_value):<15} {status:<10}")
            
            sample_count += 1
            if sample_count >= 10:  # Limit sample output
                break
        if sample_count >= 10:
            break
    
    # Final validation result
    print("\n" + "=" * 70)
    if errors == 0:
        print("✅ VALIDATION PASSED: All conversions appear to be correct!")
        success = True
    else:
        print(f"❌ VALIDATION FAILED: {errors} errors found!")
        success = False
    
    print("=" * 70)
    
    workbook.close()
    return success

if __name__ == "__main__":
    try:
        validation_success = validate_conversion()
        if validation_success:
            print("\nConversion validation completed successfully!")
        else:
            print("\nConversion validation found issues. Please review the errors above.")
        
        input("\nPress Enter to exit...")
        
    except Exception as e:
        print(f"Error during validation: {str(e)}")
        input("Press Enter to exit...")
