import openpyxl
from openpyxl import load_workbook
import re

def convert_to_number(value):
    """
    Convert a value to a number if possible, otherwise return empty string for missing data.

    Args:
        value: The cell value to convert

    Returns:
        int, float, or empty string
    """
    if value is None:
        return ""

    # Convert to string to handle various types
    str_value = str(value).strip()

    # Check for missing data indicators
    if str_value in ["--", " --", "-- ", " -- ", "", "N/A", "NA", "n/a"]:
        return ""

    # Try to convert to number
    try:
        # Remove any commas or spaces that might be in numbers
        clean_value = re.sub(r'[,\s]', '', str_value)

        # Try integer first
        if '.' not in clean_value:
            return int(float(clean_value))
        else:
            # Convert to float then to int (removes decimals)
            return int(float(clean_value))
    except (ValueError, TypeError):
        # If conversion fails, return the original value
        return value

def process_excel_file():
    """
    Process the Excel file to convert columns D to AA from row 2 onwards to numbers.
    """
    file_path = r"d:\Anant\VSCodeProjects\Temp_projects\ToNumber.xlsx"

    print("Loading Excel file...")
    workbook = load_workbook(file_path)

    # Check if RawData_Numbers sheet exists, if not copy from RawData
    if "RawData_Numbers" not in workbook.sheetnames:
        print("Creating RawData_Numbers sheet by copying RawData...")
        source_sheet = workbook["RawData"]
        target_sheet = workbook.copy_worksheet(source_sheet)
        target_sheet.title = "RawData_Numbers"
    else:
        print("Using existing RawData_Numbers sheet...")
        target_sheet = workbook["RawData_Numbers"]

    # Define the range to process (columns D to AA, rows 2 to last row)
    start_col = 4  # Column D
    end_col = 27   # Column AA
    start_row = 2

    # Find the last row with data
    last_row = target_sheet.max_row
    print(f"Processing data from row {start_row} to {last_row}, columns D to AA")

    # Process each cell in the specified range
    total_cells = (end_col - start_col + 1) * (last_row - start_row + 1)
    processed_cells = 0

    for row in range(start_row, last_row + 1):
        for col in range(start_col, end_col + 1):
            cell = target_sheet.cell(row=row, column=col)
            original_value = cell.value
            converted_value = convert_to_number(original_value)

            # Only update if the value changed
            if converted_value != original_value:
                cell.value = converted_value

            processed_cells += 1

            # Progress indicator
            if processed_cells % 1000 == 0:
                progress = (processed_cells / total_cells) * 100
                print(f"Progress: {progress:.1f}% ({processed_cells}/{total_cells} cells)")

    print("Saving the updated Excel file...")
    workbook.save(file_path)
    workbook.close()

    print("Conversion completed successfully!")
    print(f"Total cells processed: {processed_cells}")
    print(f"File saved: {file_path}")

if __name__ == "__main__":
    try:
        process_excel_file()
    except Exception as e:
        print(f"Error occurred: {str(e)}")
        input("Press Enter to exit...")
